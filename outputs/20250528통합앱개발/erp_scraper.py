"""
더존 아마란스 ERP 자동화 v2 (Playwright 기반)
- Selenium → Playwright 전환 (JS 오류 해결)
- 성화/대동 회사 전환
- 현재고현황 엑셀 → Firebase 업데이트
- 출고현황 엑셀 → 텔레그램 알림
- 생산입고현황 엑셀 → Firebase 업데이트
- 운영시간: 07:30 ~ 20:00
"""

import asyncio
import time
import json
import os
import hashlib
import glob
import requests
import openpyxl
from datetime import datetime
from playwright.async_api import async_playwright
try:
    import pyautogui
    pyautogui.FAILSAFE = False
except: pass

BASE_DIR     = os.path.dirname(os.path.abspath(__file__))
CONFIG_FILE  = os.path.join(BASE_DIR, "config.json")
DATA_FILE    = os.path.join(BASE_DIR, "last_data.json")
DOWNLOAD_DIR = os.path.join(os.path.expanduser("~"), "Downloads")
FB_URL       = "https://inventory-dce00-default-rtdb.firebaseio.com"

PRODUCTS = [
    (0,  "10000000","성화 수도용상토",               "본사"),
    (1,  "10000120","성화 청정유기능 상토",           "본사"),
    (2,  "10000121","성화 청정유기능 상토 시료",      "본사"),
    (3,  "10001000","성화 일품 상토",                "본사"),
    (4,  "10001010","성화 입상 상토",                "본사"),
    (5,  "10001060","성화 입상 상토 B(DD)",          "본사"),
    (6,  "10001070","성화 입상 상토 S(DD)",          "본사"),
    (7,  "10001080","성화 입상 상토 B",              "본사"),
    (8,  "10002050","성화 경량 플러스 상토(DD)",      "본사"),
    (9,  "10100000","성화 준중량 상토 ton bag",       "본사"),
    (10, "10100010","성화 준중량 상토 ton bag",       "본사"),
    (11, "10100050","성화 준중량유기능 상토 ton bag",  "본사"),
    (12, "10100060","성화 준중량유기능 상토 ton bag",  "본사"),
    (13, "10100360","성화 입상 상토 ton bag S(DD)",  "본사"),
    (14, "10100700","성화 일품 상토 ton bag",         "본사"),
    (15, "11000020","원예용 토백이(DD)",              "본사"),
    (16, "11000021","원예용 토백이(DD) 시료",         "본사"),
    (17, "11000050","원예용 토백이(BN)",              "본사"),
    (18, "11000051","원예용 토백이(BN) 시료",         "본사"),
    (19, "11003010","성화 딸기전용 상토(DD) 5:5",    "본사"),
    (20, "11003011","성화 딸기전용 상토(DD) 5:5 시료","본사"),
    (21, "11003030","성화 딸기전용 상토 30(DD)",      "본사"),
    (22, "12000040","바이오 미라클(DD)",              "본사"),
    (23, "12000041","바이오 미라클(DD) 시료",         "본사"),
    (24, "20000000","대동 준중량 상토나라 상토",       "본사"),
    (25, "20000001","대동 준중량 상토나라 상토 시료",  "본사"),
    (26, "20000030","대동 청정 유기농 상토",           "본사"),
    (27, "20000031","대동 청정 유기농 상토 시료",      "본사"),
    (28, "20000100","대동 경량 상토나라(DD) 상토",    "본사"),
    (29, "20000101","대동 경량 상토나라(DD) 상토 시료","본사"),
    (30, "20004000","대동 준중량 상토 ton bag",        "본사"),
    (31, "20004040","대동 준중량 유기능 상토 ton bag", "본사"),
    (32, "10000000","성화 수도용상토",                "2공장"),
    (33, "10000001","성화 수도용상토 시료",            "2공장"),
    (34, "10100000","성화 준중량 상토 ton bag",        "2공장"),
    (35, "10100360","성화 입상 상토 ton bag S(DD)",   "2공장"),
    (36, "10000000","성화 수도용상토",                "범진"),
    (37, "10000001","성화 수도용상토 시료",            "범진"),
    (38, "10000120","성화 청정유기능 상토",            "범진"),
    (39, "10001040","성화 입상 상토(DD)",             "범진"),
    (40, "10001060","성화 입상 상토 B(DD)",           "범진"),
    (41, "10001061","성화 입상 상토 B(DD) 시료",      "범진"),
    (42, "10001070","성화 입상 상토 S(DD)",           "범진"),
    (43, "10100361","성화 입상 상토 ton bag S(DD) 시료","범진"),
    (44, "30000000","풍농 엔파코 친환경 상토",         "범진"),
    (45, "10000000","성화 수도용상토",                "야적장1"),
    (46, "10000140","성화 청정 경량 유기능 상토(DD)",  "야적장1"),
    (47, "10000141","성화 청정 경량 유기능 상토(DD) 시료","야적장1"),
    (48, "10000250","성화 친환경 상토(DD)",           "야적장1"),
    (49, "10002050","성화 경량 플러스 상토(DD)",       "야적장1"),
    (50, "11000020","원예용 토백이(DD)",              "야적장1"),
    (51, "10010000","대동 준중량상토나라",             "영등"),
    (52, "10050000","대동 자연유기농상토(경량)",       "영등"),
    (53, "10120000","대동 준중량 ton bag",            "영등"),
    (54, "10210000","대동 경량 ton bag",              "영등"),
    (55, "10340000","대동 자연유기농 ton bag(경량)",   "영등"),
    (56, "11000000","원예용 상토나라",                "영등"),
    (57, "11300000","대동 딸기배지상토(5:5)",          "영등"),
    (58, "11400000","가든마켓 원예용",                "영등"),
    (59, "11710000","원예용상토(조) ton bag",          "영등"),
    (60, "13000000","어성황토",                       "영등"),
    (61, "13500000","미꼬럼방지 어성황토 ton bag",     "영등"),
    (62, "20320000","성화 입상수도용상토 B",           "영등"),
    (63, "20330000","성화 입상수도용상토 S",           "영등"),
    (64, "20331000","성화 입상(제오S)",               "영등"),
    (65, "20600000","성화 경량수도용상토",             "영등"),
    (66, "20611000","성화 경량플러스수도용상토",       "영등"),
    (67, "20700000","성화 청정경량유기능상토",         "영등"),
    (68, "20710000","성화 청정준중량유기능상토",       "영등"),
    (69, "20830000","성화 친환경 ton bag",            "영등"),
    (70, "20850000","성화 경량 ton bag",              "영등"),
    (71, "20861000","성화 경량플러스 ton bag",         "영등"),
    (72, "20862000","성화 경량플러스 ton bag",         "영등"),
    (73, "20880000","성화 청정준중량유기능 ton bag",   "영등"),
    (74, "20890000","성화 청정경량유기능 ton bag",     "영등"),
    (75, "20940000","성화 입상 B ton bag",            "영등"),
    (76, "21000000","원예용 토백이",                  "영등"),
    (77, "21060000","원예용 토백이 ton bag",           "영등"),
    (78, "21100000","양파용 토백이",                  "영등"),
    (79, "21210000","성화 딸기전용상토20",             "영등"),
    (80, "21220000","성화 딸기전용상토30",             "영등"),
    (81, "21300000","포인트블루",                     "영등"),
    (82, "21500000","블루베리배지 상토",              "영등"),
    (83, "21600000","성화 탑-스마트(양파,대파)",       "영등"),
    (84, "21700000","코코피트(100%)",                 "영등"),
    (85, "21800000","바이오차 유기농상토",             "영등"),
    (86, "23000000","바이오미라클",                   "영등"),
    (87, "23100000","조경용상토",                     "영등"),
    (88, "10010000","대동 준중량상토나라",             "야적장1(대동)"),
    (89, "10030000","대동 경량상토나라",               "야적장1(대동)"),
    (90, "32200000","한아름수도경량",                 "야적장1(대동)"),
    (91, "10030000","대동 경량상토나라",               "성화(대동)"),
]

def ts(): return datetime.now().strftime("%H:%M:%S")
def now_str(): return datetime.now().strftime("%Y-%m-%d %H:%M")

def load_config():
    with open(CONFIG_FILE, "r", encoding="utf-8") as f:
        return json.load(f)

# ── 운영시간 체크 ──────────────────────────────────────────
def is_working_hours():
    now = datetime.now()
    start = now.replace(hour=7, minute=30, second=0, microsecond=0)
    end   = now.replace(hour=20, minute=0, second=0, microsecond=0)
    return start <= now < end

def wait_until_morning(token, chat_id):
    from datetime import timedelta
    now = datetime.now()
    if now.hour >= 20:
        tomorrow = (now + timedelta(days=1)).replace(hour=7, minute=30, second=0, microsecond=0)
        wait_sec = int((tomorrow - now).total_seconds())
    else:
        today_start = now.replace(hour=7, minute=30, second=0, microsecond=0)
        wait_sec = int((today_start - now).total_seconds())
    wait_min = wait_sec // 60
    print(f"[{ts()}] 🌙 휴식 중... {wait_min}분 후 재시작 (07:30)")
    telegram(token, chat_id,
        f"🌙 <b>야간 휴식 모드</b>\n07:30에 자동 재시작합니다\n🕐 {now_str()}")
    time.sleep(wait_sec)

# ── Firebase ──────────────────────────────────────────────
def fb_get(path):
    try:
        r = requests.get(f"{FB_URL}/{path}.json", timeout=10)
        return r.json() if r.ok else None
    except Exception as e:
        print(f"[{ts()}] Firebase GET 오류: {e}"); return None

def fb_put(path, data):
    try:
        r = requests.put(f"{FB_URL}/{path}.json", json=data, timeout=10)
        return r.ok
    except Exception as e:
        print(f"[{ts()}] Firebase PUT 오류: {e}"); return False

# ── 텔레그램 ──────────────────────────────────────────────
def telegram(token, chat_id, msg):
    try:
        requests.post(f"https://api.telegram.org/bot{token}/sendMessage",
            json={"chat_id": chat_id, "text": msg, "parse_mode": "HTML"}, timeout=10)
    except Exception as e:
        print(f"[{ts()}] 텔레그램 오류: {e}")

# ── 중복 체크 ─────────────────────────────────────────────
def load_hashes():
    if not os.path.exists(DATA_FILE): return set()
    with open(DATA_FILE, "r", encoding="utf-8") as f:
        return set(json.load(f).get("hashes", []))

def save_hashes(h):
    with open(DATA_FILE, "w", encoding="utf-8") as f:
        json.dump({"hashes": list(h), "updated": datetime.now().isoformat()}, f, ensure_ascii=False)

def row_hash(r):
    return hashlib.md5(f"{r['출고일자']}|{r['고객']}|{r['품명']}|{r['규격']}|{r['단가']}".encode()).hexdigest()

# ── 엑셀 파싱: 현재고 ─────────────────────────────────────
def parse_inventory(filepath):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    header, data_start = None, 0
    for i, row in enumerate(rows):
        cells = [str(c).strip() if c else "" for c in row]
        if "품번" in cells and "기말재고" in cells:
            header = {c: j for j, c in enumerate(cells)}
            data_start = i + 1
            break
    if not header:
        print(f"[{ts()}] 현재고 헤더 못 찾음"); return {}
    code_col  = header["품번"]
    stock_col = header["기말재고"]
    loc_col   = header.get("장소") or header.get("창고") or header.get("사업장") or None
    excel_map = {}  # (code, loc) -> stock 또는 code -> stock
    for row in rows[data_start:]:
        if not any(row): continue
        code_raw  = str(row[code_col]).strip()  if row[code_col]  is not None else ""
        stock_raw = row[stock_col]
        if not code_raw or stock_raw is None or code_raw == "합계": continue
        try:
            code  = code_raw.zfill(8)
            stock = int(float(str(stock_raw).replace(",", "")))
            loc_raw = str(row[loc_col]).strip() if loc_col and row[loc_col] else ""
            key = (code, loc_raw) if loc_raw else code
            excel_map[key] = stock
        except: continue
    stock_map = {}
    for (pid, code, name, loc) in PRODUCTS:
        if (code, loc) in excel_map:
            stock_map[pid] = excel_map[(code, loc)]
        elif code in excel_map:
            stock_map[pid] = excel_map[code]
    print(f"[{ts()}] 현재고 파싱: {len(excel_map)}개 → {len(stock_map)}개 매칭 (장소컬럼: {'있음' if loc_col else '없음'})")
    return stock_map

# ── 엑셀 파싱: 출고현황 ──────────────────────────────────
def parse_shipment(filepath):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    header, data_start = None, 0
    for i, row in enumerate(rows):
        cells = [str(c).strip() if c else "" for c in row]
        if "출고일자" in cells and "고객" in cells:
            header = {c: j for j, c in enumerate(cells)}
            data_start = i + 1
            break
    if not header: return []
    def get(row, key, aliases=None):
        for k in [key] + (aliases or []):
            if k in header:
                i = header[k]
                v = row[i] if i < len(row) else None
                return str(v).strip() if v is not None else ""
        return ""
    result = []
    for row in rows[data_start:]:
        if not any(row): continue
        d, c = get(row, "출고일자"), get(row, "고객")
        if not d or not c or d == "합계": continue
        result.append({
            "출고일자": d, "고객": c,
            "품명":    get(row, "품명"),
            "규격":    get(row, "규격"),
            "출고수량": get(row, "출고수량"),
            "단가":    get(row, "단가"),
            "비고":    get(row, "비고(내역)", ["비고"]),
            "출고장소": get(row, "출고창고", ["출고장소"]),
        })
    return result

# ── 엑셀 파싱: 생산입고 ──────────────────────────────────
def parse_inbound(filepath):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    header, data_start = None, 0
    for i, row in enumerate(rows):
        cells = [str(c).strip() if c else "" for c in row]
        if "입고일자" in cells and "품명" in cells:
            header = {c: j for j, c in enumerate(cells)}
            data_start = i + 1
            break
    if not header: return []
    def get(row, key, aliases=None):
        for k in [key] + (aliases or []):
            if k in header:
                i = header[k]
                v = row[i] if i < len(row) else None
                return str(v).strip() if v is not None else ""
        return ""
    result = []
    for row in rows[data_start:]:
        if not any(row): continue
        d = get(row, "입고일자")
        if not d or d == "합계": continue
        result.append({
            "입고일자": d,
            "품번":    get(row, "품번"),
            "품명":    get(row, "품명"),
            "규격":    get(row, "규격"),
            "입고수량": get(row, "입고수량"),
            "입고장소": get(row, "입고장소", ["창고", "장소"]),
            "비고":    get(row, "비고(내역)", ["비고"]),
        })
    return result

# ── Firebase 재고 업데이트 ────────────────────────────────
def update_stocks(stock_map):
    current = fb_get("stocks")
    if not isinstance(current, list):
        current = [None] * len(PRODUCTS)
    changed = []
    for pid, val in stock_map.items():
        if pid < len(current):
            old = current[pid]
            if old != val:
                changed.append(f"{PRODUCTS[pid][2]}: {old}→{val}")
            current[pid] = val
    if fb_put("stocks", current):
        fb_put("stockMeta", {
            "updatedBy": "ERP자동",
            "updatedAt": datetime.now().strftime("%m/%d %H:%M"),
            "date": datetime.now().strftime("%Y. %m. %d.")
        })
        print(f"[{ts()}] Firebase 재고 업데이트 완료 ({len(changed)}개 변경)")
    return changed

# ── 텔레그램 알림 ─────────────────────────────────────────
def notify_shipments(new_rows, config):
    t, c = config["telegram"]["bot_token"], config["telegram"]["chat_id"]
    for r in new_rows:
        telegram(t, c,
            f"📦 <b>신규 출고 알림</b>\n"
            f"━━━━━━━━━━━━━━━━\n"
            f"📅 <b>출고일자</b>: {r['출고일자']}\n"
            f"🏢 <b>고    객</b>: {r['고객']}\n"
            f"📦 <b>품    명</b>: {r['품명']}\n"
            f"📐 <b>규    격</b>: {r['규격']}\n"
            f"🔢 <b>출고수량</b>: {r['출고수량']}\n"
            f"💰 <b>단    가</b>: {r['단가']}\n"
            f"📝 <b>비    고</b>: {r['비고'] or '-'}\n"
            f"━━━━━━━━━━━━━━━━\n"
            f"🕐 {now_str()}")
        time.sleep(0.5)

def notify_stock_update(changed, config):
    if not changed: return
    t, c = config["telegram"]["bot_token"], config["telegram"]["chat_id"]
    lines = "\n".join([f"• {x}" for x in changed[:10]])
    if len(changed) > 10: lines += f"\n... 외 {len(changed)-10}건"
    telegram(t, c,
        f"📊 <b>재고 자동 업데이트</b>\n"
        f"━━━━━━━━━━━━━━━━\n{lines}\n"
        f"━━━━━━━━━━━━━━━━\n🕐 {now_str()}")

# ── Playwright: 로그인 ────────────────────────────────────
async def pw_login(page, config):
    print(f"[{ts()}] 로그인 중...")
    await page.goto(config["erp"]["url"])
    await page.wait_for_timeout(5000)

    # 아이디 입력 - placeholder로도 찾기
    try:
        await page.wait_for_selector("#reqLoginId", state="visible", timeout=20000)
    except:
        # fallback: placeholder로 찾기
        await page.wait_for_selector("input[placeholder*='아이디']", state="visible", timeout=10000)
    
    await page.fill("input[placeholder*='아이디']", config["erp"]["username"])
    await page.wait_for_timeout(500)

    # 다음 버튼
    await page.click("button:has-text('다음')")
    await page.wait_for_timeout(4000)

    # 비밀번호 입력
    try:
        await page.wait_for_selector("#reqLoginPw", state="visible", timeout=15000)
    except:
        await page.wait_for_selector("input[placeholder*='비밀번호'], input[type='password']", state="visible", timeout=10000)
    
    await page.fill("input[type='password']", config["erp"]["password"])
    await page.wait_for_timeout(500)

    # 로그인
    await page.keyboard.press("Enter")
    await page.wait_for_timeout(7000)

    if "login" not in page.url.lower():
        print(f"[{ts()}] 로그인 완료 ✅")
        return True
    else:
        print(f"[{ts()}] 로그인 실패 ❌ URL: {page.url}")
        await page.screenshot(path=os.path.join(BASE_DIR, "login_error.png"))
        return False

# ── Playwright: 회사 전환 ─────────────────────────────────
async def pw_switch_company(page, company_text):
    """Playwright DOM 기반 회사 전환"""
    print(f"[{ts()}] 회사 전환: {company_text}")
    try:
        # 1. 우상단 프로필(송성민) 클릭
        for sel in ["[class*='userInfo']", "[class*='user-info']", "#userInfoPopupBtn"]:
            try:
                el = page.locator(sel).first
                if await el.is_visible(timeout=2000):
                    await el.click()
                    await page.wait_for_timeout(1500)
                    break
            except: pass
        else:
            await page.get_by_text("송성민", exact=False).first.click()
            await page.wait_for_timeout(1500)

        # 2. 현재 회사 텍스트 클릭 (목록 펼치기)
        for expand_text in ["(주)성화-(주)성화-영업부", "성화-영업부", "영업부"]:
            try:
                el = page.get_by_text(expand_text, exact=False).first
                if await el.is_visible(timeout=2000):
                    await el.click()
                    await page.wait_for_timeout(1500)
                    print(f"[{ts()}] 회사 목록 펼침")
                    break
            except: pass

        # 3. 목표 회사 클릭
        target_keyword = "대동산업" if "대동산업" in company_text else "성화"
        clicked = False
        for keyword in [company_text, target_keyword]:
            try:
                items = page.get_by_text(keyword, exact=False)
                count = await items.count()
                for i in range(count):
                    item = items.nth(i)
                    if await item.is_visible(timeout=1000):
                        await item.click()
                        clicked = True
                        print(f"[{ts()}] 회사 선택 클릭: {keyword}")
                        break
                if clicked:
                    break
            except: pass

        if not clicked:
            raise Exception(f"목표 회사 '{company_text}' 못 찾음")

        await page.wait_for_timeout(1000)

        # 4. 확인 팝업 버튼 클릭
        btn = page.get_by_role("button", name="확인")
        await btn.first.wait_for(state="visible", timeout=5000)
        await btn.first.click()

        await page.wait_for_timeout(5000)

        # 5. 전환 검증
        current = await page.locator(".divi_txt").first.inner_text()
        print(f"[{ts()}] 회사 전환 완료 ✅ (현재: {current.strip()})")
        return True

    except Exception as e:
        print(f"[{ts()}] 회사 전환 오류: {e}")
        await page.screenshot(path=os.path.join(BASE_DIR, "switch_error.png"))
        return False



# ── Playwright: 엑셀 다운로드 ────────────────────────────
async def pw_download_excel(page, page_url, label="", click_tab=None):
    print(f"[{ts()}] [{label}] 페이지 이동...")
    await page.goto(page_url)
    await page.wait_for_timeout(5000)

    try:
        # 공지사항 닫기
        for sel in [
            "button[class*='close']",
            "[class*='NoticeBar'] button",
            "[class*='notice'] button",
        ]:
            try:
                btn = page.locator(sel).first
                if await btn.is_visible(timeout=1000):
                    await btn.click()
                    await page.wait_for_timeout(500)
                    print(f"[{ts()}] [{label}] 공지 닫기")
            except: pass

        await page.wait_for_timeout(1000)

        # 탭 클릭 (장소/창고 탭)
        if click_tab:
            try:
                tab = page.get_by_text(click_tab, exact=True).first
                if await tab.is_visible(timeout=3000):
                    await tab.click()
                    await page.wait_for_timeout(2000)
                    print(f"[{ts()}] [{label}] 탭 클릭: {click_tab}")
            except: pass

        # JS에서 캔버스 좌표까지 한번에 가져오기 (stale 핸들 문제 방지)
        box = None
        for _ in range(15):
            box = await page.evaluate("""() => {
                const cs = Array.from(document.querySelectorAll('canvas'));
                if (!cs.length) return null;
                const largest = cs.reduce((a,b)=>(a.width*a.height)>(b.width*b.height)?a:b);
                const rect = largest.getBoundingClientRect();
                if (!rect || rect.width === 0 || rect.height === 0) return null;
                return {x: rect.x + window.scrollX, y: rect.y + window.scrollY, width: rect.width, height: rect.height};
            }""")
            if box and box['width'] > 0 and box['height'] > 0:
                break
            await page.wait_for_timeout(1000)
        if not box:
            raise Exception("Canvas bounding box 없음")

        print(f"[{ts()}] [{label}] Canvas: x={box['x']:.0f} y={box['y']:.0f} w={box['width']:.0f} h={box['height']:.0f}")

        rc_x = box['x'] + box['width'] / 2
        rc_y = box['y'] + box['height'] / 2

        # expect_download을 모든 트리거 동작 전에 감쌈 (다운로드 이벤트 누락 방지)
        async with page.expect_download(timeout=60000) as dl_info:
            # 좌클릭 후 우클릭
            await page.mouse.click(rc_x, rc_y)
            await page.wait_for_timeout(300)
            await page.mouse.click(rc_x, rc_y, button="right")
            await page.wait_for_timeout(1500)

            # 텍스트로 메뉴 항목 찾기 → 실패 시 좌표 fallback
            clicked_menu = False
            for menu_text in ["엑셀저장하기", "엑셀변환하기", "Excel", "엑셀"]:
                try:
                    item = page.get_by_text(menu_text, exact=False)
                    if await item.first.is_visible(timeout=1500):
                        await item.first.click()
                        clicked_menu = True
                        print(f"[{ts()}] [{label}] 메뉴 클릭: {menu_text}")
                        break
                except: pass

            if not clicked_menu:
                excel_x = rc_x + 60
                excel_y = rc_y - 34
                await page.mouse.click(excel_x, excel_y)
                print(f"[{ts()}] [{label}] 메뉴 좌표 클릭 fallback ({excel_x:.0f}, {excel_y:.0f})")

            await page.wait_for_timeout(1500)

            # 확인 버튼 클릭 시도 (텍스트/role로 찾기)
            confirmed = False
            for btn_text in ["확인", "저장", "다운로드", "OK"]:
                try:
                    btn = page.get_by_role("button", name=btn_text)
                    if await btn.first.is_visible(timeout=2000):
                        await btn.first.click()
                        confirmed = True
                        print(f"[{ts()}] [{label}] 확인 버튼 클릭: {btn_text}")
                        break
                except: pass

            if not confirmed:
                await page.keyboard.press("Enter")
                print(f"[{ts()}] [{label}] Enter 확인 (fallback)")

        download = await dl_info.value
        save_path = os.path.join(DOWNLOAD_DIR, download.suggested_filename)
        await download.save_as(save_path)
        print(f"[{ts()}] [{label}] 다운로드 완료: {download.suggested_filename}")
        return save_path

    except Exception as e:
        print(f"[{ts()}] [{label}] 오류: {e}")
        await page.screenshot(path=os.path.join(BASE_DIR, f"{label}_error.png"))
        return None


# ── 메인 비동기 루프 ──────────────────────────────────────
async def run_once(config):
    headless = config.get("headless", True)
    shsoil   = config["erp"].get("shsoil_company", "(주)성화 영업부(영업부)")
    daedong  = config["erp"].get("daedong_company", "(주)대동산업 영업부(영업부)")

    all_stock_map    = {}
    all_ship_rows    = []
    all_inbound_rows = []

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=headless)
        context = await browser.new_context(accept_downloads=True)
        page    = await context.new_page()

        try:
            if not await pw_login(page, config):
                raise Exception("로그인 실패")

            # 성화
            await pw_switch_company(page, shsoil)

            f = await pw_download_excel(page, config["erp"]["inventory_url"], "성화현재고", click_tab="장소")
            if f:
                all_stock_map.update(parse_inventory(f))
                try: os.remove(f)
                except: pass

            f = await pw_download_excel(page, config["erp"]["shipment_url"], "성화출고")
            if f:
                rows = parse_shipment(f)
                for r in rows: r['회사'] = '성화'
                print(f"[{ts()}] 성화 출고: {len(rows)}건")
                all_ship_rows.extend(rows)
                try: os.remove(f)
                except: pass

            f = await pw_download_excel(page, config["erp"]["inbound_url"], "성화생산입고")
            if f:
                rows = parse_inbound(f)
                for r in rows: r['회사'] = '성화'
                all_inbound_rows.extend(rows)
                try: os.remove(f)
                except: pass

            # 대동
            await pw_switch_company(page, daedong)

            f = await pw_download_excel(page, config["erp"]["inventory_url"], "대동현재고", click_tab="장소")
            if f:
                all_stock_map.update(parse_inventory(f))
                try: os.remove(f)
                except: pass

            f = await pw_download_excel(page, config["erp"]["shipment_url"], "대동출고")
            if f:
                rows = parse_shipment(f)
                for r in rows: r['회사'] = '대동산업'
                print(f"[{ts()}] 대동 출고: {len(rows)}건")
                all_ship_rows.extend(rows)
                try: os.remove(f)
                except: pass

            f = await pw_download_excel(page, config["erp"]["inbound_url"], "대동생산입고")
            if f:
                rows = parse_inbound(f)
                for r in rows: r['회사'] = '대동산업'
                all_inbound_rows.extend(rows)
                try: os.remove(f)
                except: pass

        finally:
            await browser.close()

    return all_stock_map, all_ship_rows, all_inbound_rows


# ── 메인 ─────────────────────────────────────────────────
def main():
    os.makedirs(DOWNLOAD_DIR, exist_ok=True)
    config   = load_config()
    token    = config["telegram"]["bot_token"]
    chat_id  = config["telegram"]["chat_id"]
    interval = config.get("check_interval_minutes", 30)

    print(f"\n{'='*45}")
    print(f"  더존 아마란스 ERP 자동화 v2 (Playwright)")
    print(f"  운영시간: 07:30 ~ 20:00")
    print(f"  체크 주기: {interval}분")
    print(f"{'='*45}\n")

    telegram(token, chat_id,
        f"🟢 <b>ERP 자동화 시작</b>\n"
        f"• 운영: 07:30~20:00\n"
        f"• 체크: {interval}분마다\n"
        f"🕐 {now_str()}")

    last_hashes = load_hashes()

    while True:
        if not is_working_hours() and not config.get("ignore_working_hours", False):
            wait_until_morning(token, chat_id)
            config = load_config()
            continue

        try:
            print(f"\n[{ts()}] ─── 데이터 수집 시작 ───")
            all_stock_map, all_ship_rows, all_inbound_rows = asyncio.run(run_once(config))

            if all_stock_map:
                changed = update_stocks(all_stock_map)
                notify_stock_update(changed, config)

            if all_ship_rows:
                fb_put("shipments", all_ship_rows)
                print(f"[{ts()}] Firebase shipments: {len(all_ship_rows)}건")

            if all_inbound_rows:
                fb_put("inbound", all_inbound_rows)
                print(f"[{ts()}] Firebase inbound: {len(all_inbound_rows)}건")

            cur_hashes, new_rows = set(), []
            for row in all_ship_rows:
                h = row_hash(row)
                cur_hashes.add(h)
                if h not in last_hashes:
                    new_rows.append(row)

            print(f"[{ts()}] 신규 출고: {len(new_rows)}건 / 전체: {len(all_ship_rows)}건")
            if new_rows:
                notify_shipments(new_rows, config)

            save_hashes(cur_hashes)
            last_hashes = cur_hashes

        except Exception as e:
            print(f"[{ts()}] ❌ 오류: {e}")
            telegram(token, chat_id, f"⚠️ <b>오류</b>\n{e}\n🕐 {now_str()}")

        print(f"[{ts()}] {interval}분 후 다음 체크...\n")
        time.sleep(interval * 60)


if __name__ == "__main__":
    main()
